#add 1 more frame which will get triggered once you click 'open folder 1' it will consist of 9 buttons opening specific location in storage of pc
import tkinter as tk
import time as t
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import requests
import pandas as pd
import lxml.html
import os
from openpyxl.workbook import workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
class App:



    def __init__(self, master):
        self.master = master
        self.master.title("Kurotech_scraper")

        # Create textfield and label
        self.label = tk.Label(self.master, text="Enter Company Symbol:", font=("Arial", 12))
        self.label.pack(pady=10)
        self.entry = tk.Entry(self.master, font=("Arial", 12), width=30)
        self.entry.pack(pady=10)

        # Create checkboxes
        self.var1 = tk.BooleanVar()
        self.var2 = tk.BooleanVar()
        self.check1 = tk.Checkbutton(self.master, text="Money Control", variable=self.var1, font=("Arial", 12))
        self.check2 = tk.Checkbutton(self.master, text="Screener", variable=self.var2, font=("Arial", 12))
        self.check1.pack(pady=5)
        self.check2.pack(pady=5)

        # Create button to submit choices
        self.button = tk.Button(self.master, text="Submit", font=("Arial", 12), bg="#4CAF50", fg="white",
                                command=self.submit)
        self.button.pack(pady=10)

    def submit(self):
        # Get name from text field
        name = self.entry.get()


        self.new_window1 = tk.Toplevel(self.master)
        self.new_window1.title("selected")
        self.new_window1.geometry("500x200")

        # only money control
        if self.var1.get() and not self.var2.get():
            msg = f"{name.upper()}'s data will be collected from\n MoneyControl website please press scrape to continue."
            #t.sleep(10)
            label = tk.Label(self.new_window1, text=msg, font=("Arial", 12))
            label.pack(pady=10)

            # Create buttons to initiate scrape
            self.scrape1 = tk.Button(self.new_window1, text="scrape", font=("Arial", 12), bg="#3498DB",
                                fg="white",
                                command=self.scrape1)
            self.ok = tk.Button(self.new_window1, text="ok", font=("Arial", 12), bg="#3498DB",
                                     fg="white",
                                     command=self.quit1)
            self.ok.pack(side=tk.RIGHT, padx=10, pady=10)
            self.scrape1.pack(side=tk.LEFT, padx=10, pady=10)

        #only screener
        elif self.var2.get() and not self.var1.get():
            msg = f"{name.upper()}'s data will be collected from \nScreener website please press scrape to continue."
            label = tk.Label(self.new_window1, text=msg, font=("Arial", 12))
            label.pack(pady=10)
            # Create buttons to initiate scrape
            self.scrape1 = tk.Button(self.new_window1, text="scrape", font=("Arial", 12), bg="#3498DB",
                                     fg="white",
                                     command=self.scrape2)
            self.ok = tk.Button(self.new_window1, text="ok", font=("Arial", 12), bg="#3498DB",
                                fg="white",
                                command=self.quit1)
            self.ok.pack(side=tk.RIGHT, padx=10, pady=10)
            self.scrape1.pack(side=tk.LEFT, padx=10, pady=10)

        # scrape using moneycontrol and screener
        elif self.var2.get() and self.var1.get():
            msg = f"{name}'s data will be collected from \nMoneyControl website and screener website \n\n\n\nplease press " \
                  f"scrape to continue. "
            label = tk.Label(self.new_window1, text=msg, font=("Arial", 12))
            label.pack(pady=10)
            # Create buttons to initiate scrape
            self.scrape1 = tk.Button(self.new_window1, text="scrape", font=("Arial", 12), bg="#3498DB",
                                     fg="white",
                                     command=self.scrape7)
            self.ok = tk.Button(self.new_window1, text="ok", font=("Arial", 12), bg="#3498DB",
                                fg="white",
                                command=self.quit1)
            self.ok.pack(side=tk.RIGHT, padx=10, pady=10)
            self.scrape1.pack(side=tk.LEFT, padx=10, pady=10)


        #just does nothing
        else:
            msg = f"you haven't chose any option, \nyou can look out for pre-existing files for {name} \npress continue to to analyse, open data."
            label = tk.Label(self.new_window1, text=msg, font=("Arial", 12))
            label.pack(pady=10)
            self.ok = tk.Button(self.new_window1, text="ok", font=("Arial", 12), bg="#3498DB",
                                fg="white",
                                command=self.quit1)
            self.continue1 = tk.Button(self.new_window1, text="continue", font=("Arial", 12), bg="#3498DB",
                                fg="white",
                                command=self.continue1)
            self.ok.pack(side=tk.BOTTOM, padx=10, pady=10)
            self.continue1.pack(side=tk.BOTTOM, padx=10, pady=10)











    def continue1(self):
        self.new_window2 = tk.Toplevel(self.master)
        self.new_window2.title("chose following options")
        self.new_window2.geometry("1000x200")
        label = tk.Label(self.new_window2, text="wait for the the scrape window to close", font=("Arial", 12))
        label.pack(pady=10)
        self.button1 = tk.Button(self.new_window2, text="open moneycontrol \ndata file location", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.open_MC)
        self.button2 = tk.Button(self.new_window2, text="Open screener \ndata file location", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.open_SC)
        self.button3 = tk.Button(self.new_window2, text="Analyse", font=("Arial", 12), bg="#3498DB", fg="white",
                                 command=self.analyse)
        self.button4 = tk.Button(self.new_window2, text="Update to DB", font=("Arial", 12), bg="#3498DB", fg="white",
                                 command=self.update_to_db)
        self.button5 = tk.Button(self.new_window2, text="reset", font=("Arial", 12), bg="#3498DB", fg="white",
                                 command=self.restart)
        self.button1.pack(side=tk.LEFT, padx=10, pady=10)
        self.button2.pack(side=tk.RIGHT, padx=10, pady=10)
        self.button3.pack(side=tk.TOP, padx=10, pady=10)
        self.button4.pack(side=tk.BOTTOM, padx=10, pady=10)
        self.button5.pack(side=tk.BOTTOM, padx=10, pady=10)


    def scrape1(self):
        # Open specific folder
        name = self.entry.get()
        print("starting scraping from the website")
        # for saving files in desired diretion

        os.chdir(r"D:\major_project\automated_fundamental_analysis\output_data_files\moneycontrol")
        os.mkdir(name.upper())
        os.chdir("D:\\major_project\\automated_fundamental_analysis\\output_data_files\\moneycontrol\\" + name.upper())

        path = "‪C:\\Program Files (x86)\\chromedriver.exe"
        driver = webdriver.Chrome(path)
        url = "https://www.moneycontrol.com/"
        driver.get(url)
        driver.maximize_window()
        # t.sleep(1)
        print(driver.title)
        print("_____________________________________")
        # t.sleep(1)
        search = driver.find_element(By.XPATH, '/html/body/div/div[1]/span/a')  # closing ad
        search.send_keys(Keys.ENTER)
        # t.sleep(1)
        search = driver.find_element(By.XPATH, '//*[@id="search_str"]')  # search button //*[@id="search_str"]
        search.send_keys(name)  # this one is for searching what you want
        search.send_keys(Keys.ENTER)
        # content = driver.page_source
        # print("content variable = "+content)
        # soup = BeautifulSoup(content)
        html = requests.get(driver.current_url)
        print("this is the xpath link",html)
        doc = lxml.html.fromstring(html.content)
        print("_____________________________________")
        print(doc)
        print("_____________________________________")
        balance_sheet = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[1]/a/@href')
        profit_and_loss = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[2]/a/@href')
        quarterly_result = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[3]/a/@href')
        half_yearly_result = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[4]/a/@href')
        nine_month_result = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[5]/a/@href')
        yearly_result = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[6]/a/@href')
        cashflow = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[7]/a/@href')
        ratios = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[8]/a/@href')
        capital_structure = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[9]/a/@href')
        df = pd.read_html(balance_sheet[0])
        df[0].to_excel(name + '_Balance_Sheet.xlsx')
        df = pd.read_html(profit_and_loss[0])
        df[0].to_excel(name + '_Profit_And_Loss.xlsx')
        df = pd.read_html(quarterly_result[0])
        df[0].to_excel(name + '_Quarterly_Result.xlsx')
        df = pd.read_html(half_yearly_result[0])
        df[0].to_excel(name + '_Half_Yearly_Result.xlsx')
        df = pd.read_html(nine_month_result[0])
        df[0].to_excel(name + '_Nine_Month_Result.xlsx')
        df = pd.read_html(yearly_result[0])
        df[0].to_excel(name + '_Yearly_result.xlsx')
        df = pd.read_html(cashflow[0])
        df[0].to_excel(name + '_Cashflow.xlsx')
        df = pd.read_html(ratios[0])
        df[0].to_excel(name + '_Ratios.xlsx')
        df = pd.read_html(capital_structure[0])
        df[0].to_excel(name + '_Capital_Structure.xlsx')
        # t.sleep(1)
        driver.close()
        print("execution completed code exited by [0]")
        self.new_window1.destroy()
        print("scraping ended")
        self.new_window2 = tk.Toplevel(self.master)
        self.new_window2.title("chose following options")
        self.new_window2.geometry("1000x200")
        label = tk.Label(self.new_window2, text="wait for the the scrape window to close", font=("Arial", 12))
        label.pack(pady=10)
        self.button1 = tk.Button(self.new_window2, text="open moneycontrol data file location", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.open_MC)
        self.button2 = tk.Button(self.new_window2, text="Open screener data file location", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.open_SC)
        self.button3 = tk.Button(self.new_window2, text="Analyse", font=("Arial", 12), bg="#3498DB", fg="white",
                                 command=self.analyse)
        self.button4 = tk.Button(self.new_window2, text="Update to DB", font=("Arial", 12), bg="#3498DB", fg="white",
                                 command=self.update_to_db)
        self.button5 = tk.Button(self.new_window2, text="reset", font=("Arial", 12), bg="#3498DB", fg="white",
                                 command=self.restart)
        self.button1.pack(side=tk.LEFT, padx=10, pady=10)
        self.button2.pack(side=tk.RIGHT, padx=10, pady=10)
        self.button3.pack(side=tk.TOP, padx=10, pady=10)
        self.button4.pack(side=tk.BOTTOM, padx=10, pady=10)
        self.button5.pack(side=tk.BOTTOM, padx=10, pady=10)

    def scrape2(self):
        # Open specific folder
        name = self.entry.get()


        print("screener scraping started")
        os.chdir(r"D:\major_project\automated_fundamental_analysis\output_data_files\screener")
        os.mkdir(name.upper())
        os.chdir("D:\\major_project\\automated_fundamental_analysis\\output_data_files\\screener\\" + name.upper())

        path = "‪C:\\Program Files (x86)\\chromedriver.exe"
        driver = webdriver.Chrome(path)
        url = "https://www.screener.in/"
        driver.get(url)
        driver.maximize_window()
        print(driver.title)
        print("_____________________________________")

        #                           login page button
        search = driver.find_element(By.XPATH,
                                     '/html/body/nav/div[2]/div/div/div/div[2]/div[2]/a[1]')  # this is the location of login button
        # search.send_keys(name) #this one is for searching what you want
        search.send_keys(Keys.ENTER)
        search = driver.find_element(By.XPATH,
                                     '/html/body/nav/div[2]/div/div/div/div[2]/div[2]/a[1]')  # this is the location of login button
        # search.send_keys(name) #this one is for searching what you want
        search.send_keys(Keys.ENTER)
        t.sleep(1)
        # print("_____________________________________")
        #                           login page details
        search = driver.find_element(By.XPATH, '//*[@id="id_username"]')
        search.send_keys('aditya009udemy@gmail.com')  # adityakorade009@gmail.com
        search = driver.find_element(By.XPATH, '//*[@id="id_password"]')
        search.send_keys('kuro@2247')  # kuro@2247)
        search.send_keys(Keys.ENTER)
        t.sleep(1)

        #                           searching page
        search = driver.find_element(By.XPATH,
                                     '//*[@id="desktop-search"]/div/input')  # this is the location of login button
        search.send_keys(name)  # this one is for searching what you want
        t.sleep(1)
        search.send_keys(Keys.ENTER)
        t.sleep(1)

        #                           expanding tables

        t.sleep(1)
        #                           table scraping

        new_url = driver.current_url  # getting the url of the banking statement
        print(driver.current_url)
        response = requests.get(new_url)
        print("execution completed code exited by [0]")
        df = pd.read_html(new_url)
        # print(df)
        t.sleep(1)
        df[0].to_excel(name + '_Quarterly Results.xlsx')
        df[1].to_excel(name + '_Profit & Loss.xlsx')
        df[2].to_excel(name + '_Compounded Sales Growth.xlsx')
        df[3].to_excel(name + '_Compounded Profit Growth.xlsx')
        df[4].to_excel(name + '_Stock Price CAGR.xlsx')
        df[5].to_excel(name + '_Return on Equity.xlsx')
        df[6].to_excel(name + '_Balance Sheet.xlsx')
        df[7].to_excel(name + '_Cash Flows.xlsx')
        df[8].to_excel(name + '_Cash Flows.xlsx')
        df[9].to_excel(name + '_Shareholding Pattern.xlsx')
        # name competitive company
        # //tbody/tr/td[2]/a
        print(driver.find_element(By.XPATH, '//*[@id="top-ratios"]'))
        # getting html code
        driver.close()
        print("screener scraping completed")
        self.new_window1.destroy()
        self.new_window2 = tk.Toplevel(self.master)
        self.new_window2.title("chose following options")
        self.new_window2.geometry("1000x200")
        label = tk.Label(self.new_window2, text="wait for the the scrape window to close", font=("Arial", 12))
        label.pack(pady=10)
        self.button1 = tk.Button(self.new_window2, text="open moneycontrol data file location", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.open_MC)
        self.button2 = tk.Button(self.new_window2, text="Open screener data file location", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.open_SC)
        self.button3 = tk.Button(self.new_window2, text="Analyse", font=("Arial", 12), bg="#3498DB", fg="white",
                                 command=self.analyse)
        self.button4 = tk.Button(self.new_window2, text="Update to DB", font=("Arial", 12), bg="#3498DB", fg="white",
                                 command=self.update_to_db)
        self.button5 = tk.Button(self.new_window2, text="reset", font=("Arial", 12), bg="#3498DB", fg="white",
                                 command=self.restart)
        self.button1.pack(side=tk.LEFT, padx=10, pady=10)
        self.button2.pack(side=tk.RIGHT, padx=10, pady=10)
        self.button3.pack(side=tk.TOP, padx=10, pady=10)
        self.button4.pack(side=tk.BOTTOM, padx=10, pady=10)
        self.button5.pack(side=tk.BOTTOM, padx=10, pady=10)

    def scrape7(self):
        start=t.time()
        name = self.entry.get()
        # moneycontrol(name)
        # print("_____________________________________")
        # print("enter company name")
        ##print("_____________________________________")
        print("starting scraping from the website")
        # for saving files in desired diretion

        os.chdir(r"D:\major_project\automated_fundamental_analysis\output_data_files\moneycontrol")
        os.mkdir(name.upper())
        os.chdir("D:\\major_project\\automated_fundamental_analysis\\output_data_files\\moneycontrol\\" + name.upper())

        path = "‪C:\\Program Files (x86)\\chromedriver.exe"
        driver = webdriver.Chrome(path)
        url = "https://www.moneycontrol.com/"
        driver.get(url)
        driver.maximize_window()
        # t.sleep(1)
        print(driver.title)
        print("_____________________________________")
        # t.sleep(1)
        search = driver.find_element(By.XPATH, '/html/body/div/div[1]/span/a')  # closing ad
        search.send_keys(Keys.ENTER)
        # t.sleep(1)
        search = driver.find_element(By.XPATH, '//*[@id="search_str"]')  # search button
        search.send_keys(name)  # this one is for searching what you want
        search.send_keys(Keys.ENTER)
        # content = driver.page_source
        # print("content variable = "+content)
        # soup = BeautifulSoup(content)
        html = requests.get(driver.current_url)
        print("this is the xpath link")
        doc = lxml.html.fromstring(html.content)
        print("_____________________________________")
        print(doc)
        print("_____________________________________")
        balance_sheet = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[1]/a/@href')
        profit_and_loss = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[2]/a/@href')
        quarterly_result = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[3]/a/@href')
        half_yearly_result = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[4]/a/@href')
        nine_month_result = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[5]/a/@href')
        yearly_result = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[6]/a/@href')
        cashflow = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[7]/a/@href')
        ratios = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[8]/a/@href')
        capital_structure = doc.xpath('//*[@id="consolidated"]/div[2]/div[2]/div/ul/li[9]/a/@href')
        df = pd.read_html(balance_sheet[0])
        df[0].to_excel(name + '_Balance_Sheet.xlsx')
        df = pd.read_html(profit_and_loss[0])
        df[0].to_excel(name + '_Profit_And_Loss.xlsx')
        df = pd.read_html(quarterly_result[0])
        df[0].to_excel(name + '_Quarterly_Result.xlsx')
        df = pd.read_html(half_yearly_result[0])
        df[0].to_excel(name + '_Half_Yearly_Result.xlsx')
        df = pd.read_html(nine_month_result[0])
        df[0].to_excel(name + '_Nine_Month_Result.xlsx')
        df = pd.read_html(yearly_result[0])
        df[0].to_excel(name + '_Yearly_result.xlsx')
        df = pd.read_html(cashflow[0])
        df[0].to_excel(name + '_Cashflow.xlsx')
        df = pd.read_html(ratios[0])
        df[0].to_excel(name + '_Ratios.xlsx')
        df = pd.read_html(capital_structure[0])
        df[0].to_excel(name + '_Capital_Structure.xlsx')
        # t.sleep(1)
        driver.close()
        print("execution completed code exited by [0]")

        print("scraping ended")


        print("screener scraping started")
        os.chdir(r"D:\major_project\automated_fundamental_analysis\output_data_files\screener")
        os.mkdir(name.upper())
        os.chdir("D:\\major_project\\automated_fundamental_analysis\\output_data_files\\screener\\" + name.upper())

        path = "‪C:\\Program Files (x86)\\chromedriver.exe"
        driver = webdriver.Chrome(path)
        url = "https://www.screener.in/"
        driver.get(url)
        driver.maximize_window()
        print(driver.title)
        print("_____________________________________")

        #                           login page button
        search = driver.find_element(By.XPATH,
                                     '/html/body/nav/div[2]/div/div/div/div[2]/div[2]/a[1]')  # this is the location of login button
        # search.send_keys(name) #this one is for searching what you want
        search.send_keys(Keys.ENTER)
        search = driver.find_element(By.XPATH,
                                     '/html/body/nav/div[2]/div/div/div/div[2]/div[2]/a[1]')  # this is the location of login button
        # search.send_keys(name) #this one is for searching what you want
        search.send_keys(Keys.ENTER)
        #t.sleep(1)
        # print("_____________________________________")
        #                           login page details
        search = driver.find_element(By.XPATH, '//*[@id="id_username"]')
        search.send_keys('aditya009udemy@gmail.com')  # adityakorade009@gmail.com
        search = driver.find_element(By.XPATH, '//*[@id="id_password"]')
        search.send_keys('kuro@2247')  # kuro@2247)
        search.send_keys(Keys.ENTER)
        #t.sleep(1)

        #                           searching page
        search = driver.find_element(By.XPATH,
                                     '//*[@id="desktop-search"]/div/input')  # this is the location of login button
        search.send_keys(name)  # this one is for searching what you want
        t.sleep(1)
        search.send_keys(Keys.ENTER)
        t.sleep(1)

        #                           expanding tables

        #t.sleep(1)
        #                           table scraping

        new_url = driver.current_url  # getting the url of the banking statement
        print(driver.current_url)
        response = requests.get(new_url)
        print("execution completed code exited by [0]")
        df = pd.read_html(new_url)
        # print(df)
        #t.sleep(1)
        df[0].to_excel(name + '_Quarterly Results.xlsx')
        df[1].to_excel(name + '_Profit & Loss.xlsx')
        df[2].to_excel(name + '_Compounded Sales Growth.xlsx')
        df[3].to_excel(name + '_Compounded Profit Growth.xlsx')
        df[4].to_excel(name + '_Stock Price CAGR.xlsx')
        df[5].to_excel(name + '_Return on Equity.xlsx')
        df[6].to_excel(name + '_Balance Sheet.xlsx')
        df[7].to_excel(name + '_Cash Flows.xlsx')
        df[8].to_excel(name + '_Cash Flows.xlsx')
        df[9].to_excel(name + '_Shareholding Pattern.xlsx')
        # name competitive company
        # //tbody/tr/td[2]/a
        print(driver.find_element(By.XPATH, '//*[@id="top-ratios"]'))
        # getting html code
        driver.close()
        end=t.time()
        print("time took to scrape data for one company___________________________",(start-end))
        print("screener scraping completed")
        self.new_window1.destroy()
        self.new_window2 = tk.Toplevel(self.master)
        self.new_window2.title("chose following options")
        self.new_window2.geometry("1000x200")
        label = tk.Label(self.new_window2, text="wait for the the scrape window to close", font=("Arial", 12))
        label.pack(pady=10)
        self.button1 = tk.Button(self.new_window2, text="open moneycontrol data file location", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.open_MC)
        self.button2 = tk.Button(self.new_window2, text="Open screener data file location", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.open_SC)
        self.button3 = tk.Button(self.new_window2, text="Analyse", font=("Arial", 12), bg="#3498DB", fg="white",
                                 command=self.analyse)
        self.button4 = tk.Button(self.new_window2, text="Update to DB", font=("Arial", 12), bg="#3498DB", fg="white",
                                 command=self.update_to_db)
        self.button5 = tk.Button(self.new_window2, text="reset", font=("Arial", 12), bg="#3498DB", fg="white",
                                 command=self.restart)
        self.button1.pack(side=tk.LEFT, padx=10, pady=10)
        self.button2.pack(side=tk.RIGHT, padx=10, pady=10)
        self.button3.pack(side=tk.TOP, padx=10, pady=10)
        self.button4.pack(side=tk.BOTTOM, padx=10, pady=10)
        self.button5.pack(side=tk.BOTTOM, padx=10, pady=10)

    def open_MC(self):
        # Open specific folder
        name = self.entry.get().upper()
        path_mc="D:\\major_project\\automated_fundamental_analysis\\output_data_files\\moneycontrol\\"+name
        path_mc=os.path.realpath(path_mc)
        os.startfile(path_mc)

    def open_SC(self):
        name = self.entry.get().upper()
        path_sc = "D:\\major_project\\automated_fundamental_analysis\\output_data_files\\screener\\" + name
        path_sc = os.path.realpath(path_sc)
        os.startfile(path_sc)

    def quit(self):
        # Close both windows and end program
        self.master.destroy()

    def quit1(self):
        # Close both windows and end program
        self.new_window1.destroy()

    def update_to_db(self):
        # Write code to update choices to database
        pass

    def analyse(self):
        # Open specific folder
        self.new_window3 = tk.Toplevel(self.master)
        self.new_window3.title("chose following options")
        self.new_window3.geometry("400x200")
        label = tk.Label(self.new_window3, text="wait for the the scrape window to close", font=("Arial", 12))
        label.pack(pady=10)
        self.button1 = tk.Button(self.new_window3, text="analyse company", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.solo_company)
        self.button2 = tk.Button(self.new_window3, text="compare and analyse", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.compare_company)
        self.button1.pack(side=tk.LEFT, padx=10, pady=10)
        self.button2.pack(side=tk.RIGHT, padx=10, pady=10)

    def solo_company(self):
        print("analysing just the company")
        self.new_window5 = tk.Toplevel(self.master)
        self.new_window5.title("chose following options")
        self.new_window5.geometry("400x200")
        label = tk.Label(self.new_window5, text="wait for the the scrape window to close", font=("Arial", 12))
        label.pack(pady=10)
        self.screener_s = tk.Button(self.new_window5, text="screener", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.screener_solo)
        self.moneycontrol_s = tk.Button(self.new_window5, text="moneycontrol", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.moneycontrol_solo)
        self.quit_s = tk.Button(self.new_window5, text="quit analyser", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.quit5)
        self.screener_s.pack(side=tk.LEFT, padx=10, pady=10)
        self.moneycontrol_s.pack(side=tk.RIGHT, padx=10, pady=10)
        self.quit_s.pack(side=tk.BOTTOM, padx=10, pady=10)

    def compare_company(self):
        print("comparing company")
        self.new_window4 = tk.Toplevel(self.master)
        self.new_window4.title("chose following options")
        self.new_window4.geometry("400x200")

        label = tk.Label(self.new_window4, text="enter other symbol name to compare", font=("Arial", 12))
        label.pack(pady=10)
        self.entry = tk.Entry(self.new_window4, font=("Arial", 12), width=30)
        self.entry.pack(pady=10)

        self.button1 = tk.Button(self.new_window4, text="screener", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.screener_compare)
        self.button2 = tk.Button(self.new_window4, text="moneycontrol", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.moneycontrol_compare)
        self.button3 = tk.Button(self.new_window4, text="quit analyser", font=("Arial", 12),
                                 bg="#3498DB", fg="white",
                                 command=self.quit4)
        self.button1.pack(side=tk.LEFT, padx=10, pady=10)
        self.button2.pack(side=tk.RIGHT, padx=10, pady=10)
        self.button3.pack(side=tk.BOTTOM, padx=10, pady=10)

    def quit4(self):
        self.new_window4.destroy()

    def quit5(self):
        self.new_window5.destroy()

    def moneycontrol_compare(self):
        print("comparing files from moneycontrol")

    def screener_compare(self):
        print("comparing files from screener")

    def screener_solo(self):
        print("analysing files from screener")
        name = self.entry.get()
        print(name)
        excelfile = 'D:\\major_project\\automated_fundamental_analysis\\output_data_files\\moneycontrol\\' + name.upper() + "\\" + name + '_Ratios.xlsx'
        wb = load_workbook(excelfile)
        ws = wb.active
        years = [ws['G2'].value, ws['F2'].value, ws['E2'].value, ws['D2'].value, ws['C2'].value]

        # Per Share Ratios
        figure, axis = plt.subplots(2, 3)
        figure.suptitle(ws['B3'].value)
        value = [float(ws['G9'].value), float(ws['F9'].value), float(ws['E9'].value), float(ws['D9'].value),
                 float(ws['C9'].value)]
        axis[0, 0].plot(years, value)
        axis[0, 0].plot(years, value)
        axis[0, 0].set_title("dividend")

        value = [float(ws['G10'].value), float(ws['F10'].value), float(ws['E10'].value), float(ws['D10'].value),
                 float(ws['C10'].value)]
        axis[0, 1].plot(years, value)
        axis[0, 1].plot(years, value)
        axis[0, 1].set_title(ws['B10'].value)

        value = [float(ws['G11'].value), float(ws['F11'].value), float(ws['E11'].value), float(ws['D11'].value),
                 float(ws['C11'].value)]
        axis[0, 2].plot(years, value)
        axis[0, 2].plot(years, value)
        axis[0, 2].set_title(ws['B11'].value)

        value = [float(ws['G12'].value), float(ws['F12'].value), float(ws['E12'].value), float(ws['D12'].value),
                 float(ws['C12'].value)]
        axis[1, 0].plot(years, value)

        axis[1, 0].plot(years, value)
        axis[1, 0].set_title(ws['B12'].value)

        value = [float(ws['G13'].value), float(ws['F13'].value), float(ws['E13'].value), float(ws['D13'].value),
                 float(ws['C13'].value)]
        axis[1, 1].plot(years, value)

        axis[1, 1].plot(years, value)
        axis[1, 1].set_title(ws['B13'].value)

        value = [float(ws['G14'].value), float(ws['F14'].value), float(ws['E14'].value), float(ws['D14'].value),
                 float(ws['C14'].value)]
        axis[1, 2].plot(years, value)
        axis[1, 2].plot(years, value)
        axis[1, 2].set_title(ws['B14'].value)
        plt.show()

        # Profitability Ratios
        figure1, axis1 = plt.subplots(2, 3)
        figure1.suptitle(ws['B15'].value)
        value = [float(ws['G19'].value), float(ws['F19'].value), float(ws['E19'].value), float(ws['D19'].value),
                 float(ws['C19'].value)]
        axis1[0, 0].plot(years, value)
        axis1[0, 0].plot(years, value)
        axis1[0, 0].set_title(ws['B19'].value)

        value = [float(ws['G20'].value), float(ws['F20'].value), float(ws['E20'].value), float(ws['D20'].value),
                 float(ws['C20'].value)]
        axis1[0, 1].plot(years, value)
        axis1[0, 1].plot(years, value)
        axis1[0, 1].set_title(ws['B20'].value)

        value = [float(ws['G21'].value), float(ws['F21'].value), float(ws['E21'].value), float(ws['D21'].value),
                 float(ws['C21'].value)]
        axis1[0, 2].plot(years, value)
        axis1[0, 2].plot(years, value)
        axis1[0, 2].set_title(ws['B21'].value)

        value = [float(ws['G22'].value), float(ws['F22'].value), float(ws['E22'].value), float(ws['D22'].value),
                 float(ws['C22'].value)]
        axis1[1, 0].plot(years, value)
        axis1[1, 0].plot(years, value)
        axis1[1, 0].set_title(ws['B22'].value)

        value = [float(ws['G23'].value), float(ws['F23'].value), float(ws['E23'].value), float(ws['D23'].value),
                 float(ws['C23'].value)]

        axis1[1, 1].plot(years, value)
        axis1[1, 1].set_title(ws['B23'].value)

        value = [float(ws['G24'].value), float(ws['F24'].value), float(ws['E24'].value), float(ws['D24'].value),
                 float(ws['C24'].value)]
        axis1[1, 2].plot(years, value)
        axis1[1, 2].plot(years, value)
        axis1[1, 2].set_title(ws['B24'].value)
        plt.show()

        years = [ws['G2'].value, ws['F2'].value, ws['E2'].value, ws['D2'].value, ws['C2'].value]

        # Liquidity Ratios
        figure1, axis1 = plt.subplots(2, 3)
        figure1.suptitle(ws['B25'].value + " & " + ws['B33'].value)
        value = [float(ws['G36'].value), float(ws['F36'].value), float(ws['E36'].value), float(ws['D36'].value),
                 float(ws['C36'].value)]
        axis1[0, 0].plot(years, value)
        axis1[0, 0].plot(years, value)
        axis1[0, 0].set_title(ws['B36'].value)

        value = [float(ws['G38'].value), float(ws['F38'].value), float(ws['E38'].value), float(ws['D38'].value),
                 float(ws['C38'].value)]
        axis1[0, 1].plot(years, value)
        axis1[0, 1].set_title(ws['B38'].value)

        value = [float(ws['G39'].value), float(ws['F39'].value), float(ws['E39'].value), float(ws['D39'].value),
                 float(ws['C39'].value)]
        axis1[0, 2].plot(years, value)
        axis1[0, 2].set_title(ws['B39'].value)

        plt.show()

    def moneycontrol_solo(self):
        print("comparing files from moneycontrol")
        name = self.entry.get()
        print(name)
        excelfile = 'D:\\major_project\\automated_fundamental_analysis\\output_data_files\\moneycontrol\\' + name.upper() + "\\" + name + '_Ratios.xlsx'
        wb = load_workbook(excelfile)
        ws = wb.active
        years = [ws['G2'].value, ws['F2'].value, ws['E2'].value, ws['D2'].value, ws['C2'].value]

        # Per Share Ratios
        figure, axis = plt.subplots(2, 3)
        figure.suptitle(ws['B3'].value)
        value = [float(ws['G9'].value), float(ws['F9'].value), float(ws['E9'].value), float(ws['D9'].value),
                 float(ws['C9'].value)]
        axis[0, 0].plot(years, value)
        axis[0, 0].plot(years, value)
        axis[0, 0].set_title("dividend")

        value = [float(ws['G10'].value), float(ws['F10'].value), float(ws['E10'].value), float(ws['D10'].value),
                 float(ws['C10'].value)]
        axis[0, 1].plot(years, value)
        axis[0, 1].plot(years, value)
        axis[0, 1].set_title(ws['B10'].value)

        value = [float(ws['G11'].value), float(ws['F11'].value), float(ws['E11'].value), float(ws['D11'].value),
                 float(ws['C11'].value)]
        axis[0, 2].plot(years, value)
        axis[0, 2].plot(years, value)
        axis[0, 2].set_title(ws['B11'].value)

        value = [float(ws['G12'].value), float(ws['F12'].value), float(ws['E12'].value), float(ws['D12'].value),
                 float(ws['C12'].value)]
        axis[1, 0].plot(years, value)

        axis[1, 0].plot(years, value)
        axis[1, 0].set_title(ws['B12'].value)

        value = [float(ws['G13'].value), float(ws['F13'].value), float(ws['E13'].value), float(ws['D13'].value),
                 float(ws['C13'].value)]
        axis[1, 1].plot(years, value)

        axis[1, 1].plot(years, value)
        axis[1, 1].set_title(ws['B13'].value)

        value = [float(ws['G14'].value), float(ws['F14'].value), float(ws['E14'].value), float(ws['D14'].value),
                 float(ws['C14'].value)]
        axis[1, 2].plot(years, value)
        axis[1, 2].plot(years, value)
        axis[1, 2].set_title(ws['B14'].value)
        plt.show()

        # Profitability Ratios
        figure1, axis1 = plt.subplots(2, 3)
        figure1.suptitle(ws['B15'].value)
        value = [float(ws['G19'].value), float(ws['F19'].value), float(ws['E19'].value), float(ws['D19'].value),
                 float(ws['C19'].value)]
        axis1[0, 0].plot(years, value)
        axis1[0, 0].plot(years, value)
        axis1[0, 0].set_title(ws['B19'].value)

        value = [float(ws['G20'].value), float(ws['F20'].value), float(ws['E20'].value), float(ws['D20'].value),
                 float(ws['C20'].value)]
        axis1[0, 1].plot(years, value)
        axis1[0, 1].plot(years, value)
        axis1[0, 1].set_title(ws['B20'].value)

        value = [float(ws['G21'].value), float(ws['F21'].value), float(ws['E21'].value), float(ws['D21'].value),
                 float(ws['C21'].value)]
        axis1[0, 2].plot(years, value)
        axis1[0, 2].plot(years, value)
        axis1[0, 2].set_title(ws['B21'].value)

        value = [float(ws['G22'].value), float(ws['F22'].value), float(ws['E22'].value), float(ws['D22'].value),
                 float(ws['C22'].value)]
        axis1[1, 0].plot(years, value)
        axis1[1, 0].plot(years, value)
        axis1[1, 0].set_title(ws['B22'].value)

        value = [float(ws['G23'].value), float(ws['F23'].value), float(ws['E23'].value), float(ws['D23'].value),
                 float(ws['C23'].value)]

        axis1[1, 1].plot(years, value)
        axis1[1, 1].set_title(ws['B23'].value)

        value = [float(ws['G24'].value), float(ws['F24'].value), float(ws['E24'].value), float(ws['D24'].value),
                 float(ws['C24'].value)]
        axis1[1, 2].plot(years, value)
        axis1[1, 2].plot(years, value)
        axis1[1, 2].set_title(ws['B24'].value)
        plt.show()

        years = [ws['G2'].value, ws['F2'].value, ws['E2'].value, ws['D2'].value, ws['C2'].value]

        # Liquidity Ratios
        figure1, axis1 = plt.subplots(2, 3)
        figure1.suptitle(ws['B25'].value + " & " + ws['B33'].value)
        value = [float(ws['G36'].value), float(ws['F36'].value), float(ws['E36'].value), float(ws['D36'].value),
                 float(ws['C36'].value)]
        axis1[0, 0].plot(years, value)
        axis1[0, 0].plot(years, value)
        axis1[0, 0].set_title(ws['B36'].value)

        value = [float(ws['G38'].value), float(ws['F38'].value), float(ws['E38'].value), float(ws['D38'].value),
                 float(ws['C38'].value)]
        axis1[0, 1].plot(years, value)
        axis1[0, 1].set_title(ws['B38'].value)

        value = [float(ws['G39'].value), float(ws['F39'].value), float(ws['E39'].value), float(ws['D39'].value),
                 float(ws['C39'].value)]
        axis1[0, 2].plot(years, value)
        axis1[0, 2].set_title(ws['B39'].value)

        plt.show()

    def restart(self):
        # Close choices window and reset text field and checkboxes
        self.new_window2.destroy()
        self.new_window1.destroy()
        self.new_window3.destroy()
        self.entry.delete(0, tk.END)
        self.var1.set(False)
        self.var2.set(False)




# Create main window and run app
root = tk.Tk()
root.geometry("400x300")
app = App(root)
root.mainloop()

